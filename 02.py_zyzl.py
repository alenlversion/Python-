import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
from datetime import datetime, timedelta

# ================== 配置区 ==================
INPUT_FILE = "data/科室数据汇总-2024年.xlsx"  # 您上传的文件
OUTPUT_FILE = "data/科室诊疗量对比分析_专业报告.xlsx"

# ============================================

def excel_date_to_month(excel_date_num):
    """
    将Excel日期序列号转换为"YYYY年M月"格式
    45658 -> 2025-01-01 -> "2025年1月"
    """
    try:
        # Excel日期基准：1900-01-01 = 1
        base_date = datetime(1899, 12, 30)  # Excel的日期基准
        actual_date = base_date + timedelta(days=float(excel_date_num))
        return f"{actual_date.year}年{actual_date.month}月"
    except:
        return str(excel_date_num)


def load_sheet_data(sheet_idx, year_label):
    """加载并标准化指定工作表的数据"""
    try:
        # 读取原始数据（不设置header，以便处理列标题）
        df_raw = pd.read_excel(INPUT_FILE, sheet_name=sheet_idx, header=None)
    except Exception as e:
        raise ValueError(f"读取工作表失败（索引{sheet_idx}）：{str(e)}")

    # 识别标题行（通常第0行包含"科室"）
    header_row = 0
    for i, val in enumerate(df_raw.iloc[:, 0]):
        if str(val).strip() == "科室":
            header_row = i
            break

    # 重新设置列名（使用header_row+1行）
    df = df_raw.iloc[header_row + 1:].copy()
    df.columns = df_raw.iloc[header_row].tolist()

    # 清洗科室列
    dept_col = None
    for col in df.columns:
        if "科室" in str(col):
            dept_col = col
            break
    if dept_col is None:
        raise ValueError("未找到科室列！")

    # 重置索引为科室名
    df = df.set_index(dept_col)
    df.index.name = "科室"
    df.index = df.index.astype(str).str.strip()

    # 清洗特殊字符
    df.index = df.index.str.replace(r"[^\w\u4e00-\u9fa5\-]", "", regex=True)
    df.index = df.index.str.replace(r"\s+", "", regex=True)

    # 过滤合计/总计行
    df = df[~df.index.str.contains(r"^(?:合计|总计|小计|总|全年)$", na=False, regex=True)]
    df = df[df.index != ""]

    # 转换数值
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # 处理列名：转换Excel日期序列号 -> 月份
    new_columns = []
    for col in df.columns:
        col_str = str(col).strip()
        # 尝试转换Excel日期序列号
        try:
            col_num = float(col_str)
            if 40000 < col_num < 60000:  # 合理的Excel日期范围
                new_name = excel_date_to_month(col_num)
                new_columns.append(new_name)
                continue
        except:
            pass

        # 尝试识别年份+月份
        match = re.search(r"(202[45])年?(\d{1,2})月?", col_str)
        if match:
            year = match.group(1)
            month = int(match.group(2))
            new_columns.append(f"{year}年{month}月")
            continue

        # 识别"汇总"列
        if any(kw in col_str.lower() for kw in ["汇总", "总计", "合计", "总"]):
            new_columns.append(f"{year_label}年汇总")
            continue

        new_columns.append(col_str)

    df.columns = new_columns

    # 如果没有明确的汇总列，自动计算
    if f"{year_label}年汇总" not in df.columns:
        month_cols = [c for c in df.columns if re.search(r"\d{4}年\d{1,2}月", c)]
        if month_cols:
            df[f"{year_label}年汇总"] = df[month_cols].sum(axis=1)

    return df


def safe_growth_rate(current, base):
    """安全计算增长率"""
    if base == 0 and current == 0:
        return "-"
    if base == 0:
        return "∞" if current > 0 else "-∞"
    return round((current - base) / base * 100, 2)


def main():
    print("=" * 60)
    print("🚀 科室诊疗量年度对比分析系统 v3.0")
    print(f"📁 源文件: {INPUT_FILE}")
    print("=" * 60)

    # 检查文件是否存在
    if not os.path.exists(INPUT_FILE):
        print(f"\n❌ 源文件不存在: '{INPUT_FILE}'")
        print("   请确认文件在当前目录，或修改脚本中的INPUT_FILE配置")
        return

    # 加载数据
    print("\n🔍 正在加载2024年和2025年数据...")
    try:
        df_25 = load_sheet_data(0, "2025")  # Sheet1 = 2025年
        df_24 = load_sheet_data(1, "2024")  # Sheet2 = 2024年
        print(f"✅ 数据加载成功！")
        print(f"   • 2025年科室数: {len(df_25)}")
        print(f"   • 2024年科室数: {len(df_24)}")
        print(f"   • 共同科室数: {len(set(df_24.index) & set(df_25.index))}")
    except Exception as e:
        print(f"\n❌ 数据加载失败: {str(e)}")
        print("   请检查Excel文件结构是否符合预期")
        import traceback
        traceback.print_exc()
        return

    # 合并数据
    df_merged = df_24.join(df_25, how='outer').fillna(0)

    # 获取标准月份列表
    months_std = []
    for year in ["2024", "2025"]:
        for month in range(1, 13):
            months_std.append(f"{year}年{month}月")

    # ========== 1. 原始数据工作表 ==========
    df_raw = df_merged.reset_index()

    # ========== 2. 科室汇总对比 ==========
    summary_list = []
    for dept in df_merged.index:
        # 获取2024年汇总
        if "2024年汇总" in df_merged.columns:
            v24 = df_merged.loc[dept, "2024年汇总"]
        else:
            cols_24 = [c for c in df_merged.columns if c.startswith("2024年") and "月" in c]
            v24 = df_merged.loc[dept, cols_24].sum() if cols_24 else 0

        # 获取2025年汇总
        if "2025年汇总" in df_merged.columns:
            v25 = df_merged.loc[dept, "2025年汇总"]
        else:
            cols_25 = [c for c in df_merged.columns if c.startswith("2025年") and "月" in c]
            v25 = df_merged.loc[dept, cols_25].sum() if cols_25 else 0

        bias_val = v25 - v24
        growth_rate = safe_growth_rate(v25, v24)

        # 智能偏差方向标注
        if v24 == 0 and v25 == 0:
            direction = "⚠️ 无业务记录"
        elif growth_rate == "∞":
            direction = "🚀 新增业务↑"
        elif growth_rate == "-∞":
            direction = "📉 业务归零↓"
        elif growth_rate == "-":
            direction = "→ 无变化"
        elif isinstance(growth_rate, (int, float)):
            if abs(growth_rate) >= 30:
                direction = "↑显著增长" if growth_rate > 0 else "↓大幅下降"
            elif abs(growth_rate) < 5:
                direction = "→基本持平"
            else:
                direction = "↑小幅增长" if growth_rate > 0 else "↓小幅下降"
        else:
            direction = "❓异常"

        summary_list.append({
            '科室': dept,
            '2024年汇总': v24,
            '2025年汇总': v25,
            '偏差值': bias_val,
            '增长率(%)': growth_rate,
            '业务状态': direction
        })

    df_summary = pd.DataFrame(summary_list)
    df_summary = df_summary.sort_values('偏差值', ascending=False).reset_index(drop=True)

    # ========== 3. 月度详细对比 ==========
    monthly_list = []
    for dept in df_merged.index:
        for month_num in range(1, 13):
            col24 = f"2024年{month_num}月"
            col25 = f"2025年{month_num}月"

            v24 = df_merged.loc[dept, col24] if col24 in df_merged.columns else 0
            v25 = df_merged.loc[dept, col25] if col25 in df_merged.columns else 0

            monthly_list.append({
                '科室': dept,
                '月份': f"{month_num}月",
                '2024年值': v24,
                '2025年值': v25,
                '偏差值': v25 - v24,
                '增长率(%)': safe_growth_rate(v25, v24)
            })

    df_monthly = pd.DataFrame(monthly_list)

    # ========== 4. 全院月度总览 ==========
    total_list = []
    total_24_sum = 0
    total_25_sum = 0

    for month_num in range(1, 13):
        col24 = f"2024年{month_num}月"
        col25 = f"2025年{month_num}月"

        v24 = df_merged[col24].sum() if col24 in df_merged.columns else 0
        v25 = df_merged[col25].sum() if col25 in df_merged.columns else 0

        total_24_sum += v24
        total_25_sum += v25

        # 计算环比（与上月比）
        if month_num == 1:
            mom = "-"
        else:
            prev_col25 = f"2025年{month_num - 1}月"
            prev_v25 = df_merged[prev_col25].sum() if prev_col25 in df_merged.columns else 0
            mom = safe_growth_rate(v25, prev_v25)
            if mom == "-":
                mom = "-"
            elif mom == "∞" or mom == "-∞":
                mom = mom
            else:
                mom = f"{mom}%"

        total_list.append({
            '月份': f"{month_num}月",
            '2024年总和': v24,
            '2025年总和': v25,
            '偏差值': v25 - v24,
            '增长率(%)': safe_growth_rate(v25, v24),
            '2025年环比(%)': mom
        })

    # 全年合计
    total_list.append({
        '月份': '全年合计',
        '2024年总和': total_24_sum,
        '2025年总和': total_25_sum,
        '偏差值': total_25_sum - total_24_sum,
        '增长率(%)': safe_growth_rate(total_25_sum, total_24_sum),
        '2025年环比(%)': "-"
    })

    df_total = pd.DataFrame(total_list)

    # ========== 5. 纵向对比含增长率（完美匹配您要求的样式）==========
    long_list = []
    for dept in df_merged.index:
        # 2024年行
        row_24 = {'科室': dept, '年份': 2024}
        for month_num in range(1, 13):
            col = f"2024年{month_num}月"
            row_24[f"{month_num}月"] = df_merged.loc[dept, col] if col in df_merged.columns else 0

        # 计算2024年汇总
        if "2024年汇总" in df_merged.columns:
            row_24['汇总'] = df_merged.loc[dept, "2024年汇总"]
        else:
            month_cols = [f"2024年{m}月" for m in range(1, 13)]
            row_24['汇总'] = sum(row_24.get(f"{m}月", 0) for m in range(1, 13))

        row_24['较2024年增长率(%)'] = ""  # 2024年行留空
        long_list.append(row_24)

        # 2025年行
        row_25 = {'科室': dept, '年份': 2025}
        for month_num in range(1, 13):
            col = f"2025年{month_num}月"
            row_25[f"{month_num}月"] = df_merged.loc[dept, col] if col in df_merged.columns else 0

        # 计算2025年汇总
        if "2025年汇总" in df_merged.columns:
            row_25['汇总'] = df_merged.loc[dept, "2025年汇总"]
        else:
            month_cols = [f"2025年{m}月" for m in range(1, 13)]
            row_25['汇总'] = sum(row_25.get(f"{m}月", 0) for m in range(1, 13))

        # 计算增长率
        row_25['较2024年增长率(%)'] = safe_growth_rate(row_25['汇总'], row_24['汇总'])
        long_list.append(row_25)

    # 创建DataFrame并排序
    df_long = pd.DataFrame(long_list)
    month_cols = [f"{m}月" for m in range(1, 13)]
    cols_order = ['科室', '年份'] + month_cols + ['汇总', '较2024年增长率(%)']
    df_long = df_long[cols_order].sort_values(['科室', '年份']).reset_index(drop=True)

    # ========== 6. 生成Excel报告 ==========
    print("\n💾 正在生成专业分析报告...")
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df_raw.to_excel(writer, sheet_name='1_原始数据', index=False)
        df_summary.to_excel(writer, sheet_name='2_科室汇总对比', index=False)
        df_monthly.to_excel(writer, sheet_name='3_月度详细对比', index=False)
        df_total.to_excel(writer, sheet_name='4_全院月度总览', index=False)
        df_long.to_excel(writer, sheet_name='5_纵向对比含增长率', index=False)

    # ========== 7. 专业美化 ==========
    wb = writer.book

    # 样式定义
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # 深蓝
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    neg_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 负值红
    pos_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 正值绿
    inf_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # ∞蓝色
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")  # 合计浅蓝
    alert_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 警告黄

    # 需要特殊处理的列
    RATE_COLS = ["增长率(%)", "较2024年增长率(%)", "2025年环比(%)"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 获取列名
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        # 需要特殊处理的列索引
        rate_col_indices = []
        for col_name in RATE_COLS:
            if col_name in headers:
                rate_col_indices.append(headers.index(col_name) + 1)  # openpyxl列索引从1开始

        # 遍历所有单元格
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                # 应用边框
                cell.border = border

                if row_idx == 1:  # 标题行
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # 数据行
                    # 特殊列：增长率/偏差率
                    if col_idx in rate_col_indices:
                        val_str = str(cell.value).strip() if cell.value else ""

                        # 处理特殊值
                        if val_str in ["∞", "-∞"]:
                            cell.fill = inf_fill
                        elif val_str == "-":
                            cell.fill = alert_fill
                        else:
                            try:
                                num_val = float(val_str)
                                if num_val < 0:
                                    cell.fill = neg_fill
                                elif num_val > 0:
                                    cell.fill = pos_fill
                            except (ValueError, TypeError):
                                pass

                    # 全年合计行
                    if sheet_name == '4_全院月度总览' and row_idx == ws.max_row:
                        cell.font = Font(bold=True)
                        cell.fill = total_fill

                    # 无业务记录科室
                    if sheet_name == '2_科室汇总对比':
                        status_col = headers.index("业务状态") + 1 if "业务状态" in headers else None
                        if status_col and col_idx == status_col and "无业务记录" in str(cell.value):
                            cell.fill = alert_fill

                    # 纵向对比中的2024年行
                    if sheet_name == '5_纵向对比含增长率':
                        year_col = headers.index("年份") + 1 if "年份" in headers else None
                        if year_col and col_idx == year_col and cell.value == 2024:
                            for c in row:
                                c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # 自适应列宽
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)

            for cell in col:
                if cell.value:
                    # 处理特殊值显示
                    if str(cell.value) in ["∞", "-∞"]:
                        cell.value = str(cell.value) + " (新增/归零)"

                    cell_len = len(str(cell.value))
                    max_len = max(max_len, cell_len)

            # 设置列宽（最小10，最大30）
            adjusted_width = min(max(12, max_len + 2), 30)
            ws.column_dimensions[col_letter].width = adjusted_width

    # 保存
    wb.save(OUTPUT_FILE)

    # ========== 8. 生成报告 ==========
    total_24 = df_summary['2024年汇总'].sum()
    total_25 = df_summary['2025年汇总'].sum()
    overall_growth = safe_growth_rate(total_25, total_24)

    # 统计特殊科室
    zero_depts = df_summary[df_summary['2024年汇总'] + df_summary['2025年汇总'] == 0]['科室'].tolist()
    new_depts = df_summary[df_summary['增长率(%)'] == "∞"]['科室'].tolist()
    drop_depts = df_summary[
        (df_summary['增长率(%)'].apply(lambda x: isinstance(x, float) and x < -30)) |
        (df_summary['增长率(%)'] == "-∞")
        ]['科室'].tolist()

    print("\n" + "=" * 70)
    print(f"✅ 专业分析报告生成成功: {OUTPUT_FILE}")
    print("=" * 70)
    print(f"📊 全局数据概览")
    print(f"   • 2024年总诊疗量: {int(total_24):,} 例")
    print(f"   • 2025年总诊疗量: {int(total_25):,} 例")
    print(f"   • 全年净变化: {int(total_25 - total_24):,} 例 ({overall_growth}%)")
    print(f"\n🔍 重点科室分析")
    print(
        f"   • 无业务记录科室: {len(zero_depts)} 个 (如: {', '.join(zero_depts[:3]) + '...' if len(zero_depts) > 3 else ', '.join(zero_depts) or '无'})")
    print(
        f"   • 新增业务科室: {len(new_depts)} 个 (如: {', '.join(new_depts[:3]) + '...' if len(new_depts) > 3 else ', '.join(new_depts) or '无'})")
    print(
        f"   • 大幅下降科室(↓>30%): {len(drop_depts)} 个 (如: {', '.join(drop_depts[:3]) + '...' if len(drop_depts) > 3 else ', '.join(drop_depts) or '无'})")
    print(f"\n📁 报告包含5个工作表:")
    print("   [1_原始数据]           → 原始数据对照")
    print("   [2_科室汇总对比]       → 含业务状态智能标注（⚠️无业务/🚀新增/📉归零）")
    print("   [3_月度详细对比]       → 各科室每月明细（含增长率）")
    print("   [4_全院月度总览]       → 全院月度趋势+环比分析（全年合计高亮）")
    print("   [5_纵向对比含增长率]   → 完美匹配您要求的表格样式！")
    print(f"\n💡 使用建议:")
    print("   • 向领导汇报：首选【5_纵向对比含增长率】工作表")
    print("   • 业务分析：在【2_科室汇总对比】筛选【业务状态】列")
    print("   • 月度复盘：查看【4_全院月度总览】中的环比数据")
    print("=" * 70)
    print(f"✨ 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("📌 注: ∞ = 2024年为0但2025年有值（新增业务） | -∞ = 业务归零")
    print("=" * 70)


if __name__ == "__main__":
    # 依赖检查
    try:
        import pandas
        import numpy
        import openpyxl
    except ImportError as e:
        print(f"❌ 依赖缺失: {str(e)}")
        print("   请安装所需库: pip install pandas numpy openpyxl")
        exit(1)

    # 检查文件
    if not os.path.exists(INPUT_FILE):
        print(f"\n❌ 源文件不存在: '{INPUT_FILE}'")
        print("   请将您上传的Excel文件放在当前目录")
        exit(1)

    # 执行主函数
    try:
        main()
    except PermissionError:
        print(f"\n❌ 文件权限错误！请关闭 '{OUTPUT_FILE}' 或 '{INPUT_FILE}' 后重试")
        exit(1)
    except Exception as e:
        print(f"\n❌ 处理失败: {type(e).__name__}: {str(e)}")
        print("🔍 调试建议:")
        print("  1. 确认Excel文件未被其他程序打开")
        print("  2. 检查工作表结构是否符合预期")
        print("  3. 如问题持续，请提供错误详情")
        import traceback

        traceback.print_exc()
        exit(1)