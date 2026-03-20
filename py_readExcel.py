import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Color


def generate_supermarket_pro(filename):
    wb = openpyxl.Workbook()

    # --- 1. Sheet4: 进货单价 (成本) ---
    ws_cost = wb.active
    ws_cost.title = "进货单价"
    ws_cost.append(["商品名称", "进货成本(元)"])
    costs = [["苹果", 3.5], ["洗发水", 25.0], ["方便面", 2.0], ["鲈鱼", 15.0]]
    for row in costs:
        ws_cost.append(row)

    # --- 2. Sheet3: 售卖金额 (零售价) ---
    ws_price = wb.create_sheet("售卖金额")
    ws_price.append(["商品名称", "零售单价(元)"])
    prices = [["苹果", 6.5], ["洗发水", 48.0], ["方便面", 4.0], ["鲈鱼", 32.0]]
    for row in prices:
        ws_price.append(row)

    # --- 3. Sheet2: 每日销量 ---
    ws_qty = wb.create_sheet("每日销量")
    ws_qty.append(["日期", "商品名称", "销售数量"])
    sales_records = [
        ["01-01", "苹果", 50], ["01-01", "洗发水", 5],
        ["01-02", "方便面", 100], ["01-02", "鲈鱼", 10],
        ["01-03", "苹果", 40], ["01-03", "洗发水", 8],
        ["01-04", "方便面", 80], ["01-04", "鲈鱼", 15],
    ]
    for row in sales_records:
        ws_qty.append(row)

    # --- 4. Sheet1: 管理看板 (深度分析) ---
    ws_dash = wb.create_sheet("管理看板", 0)

    # 标题
    ws_dash["A1"] = "超市月度盈利深度分析看板"
    ws_dash.merge_cells("A1:G1")
    ws_dash["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws_dash["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_dash["A1"].alignment = Alignment(horizontal="center")

    # A. 核心数据透视汇总 (使用复杂的跨表 VLOOKUP/SUMIF 公式)
    ws_dash["A3"] = "商品名称"
    ws_dash["B3"] = "总销量"
    ws_dash["C3"] = "总收入"
    ws_dash["D3"] = "总成本"
    ws_dash["E3"] = "毛利润"

    products = ["苹果", "洗发水", "方便面", "鲈鱼"]
    for i, p in enumerate(products, start=4):
        ws_dash.cell(row=i, column=1, value=p)
        # 总销量：从 Sheet2 统计
        ws_dash.cell(row=i, column=2, value=f"=SUMIF('每日销量'!B:B, A{i}, '每日销量'!C:C)")
        # 总收入：销量 * Sheet3 的单价
        ws_dash.cell(row=i, column=3, value=f"=B{i} * VLOOKUP(A{i}, '售卖金额'!A:B, 2, FALSE)")
        # 总成本：销量 * Sheet4 的成本
        ws_dash.cell(row=i, column=4, value=f"=B{i} * VLOOKUP(A{i}, '进货单价'!A:B, 2, FALSE)")
        # 毛利润：收入 - 成本
        ws_dash.cell(row=i, column=5, value=f"=C{i}-D{i}")

    # B. 总体经营结论
    ws_dash["G4"] = "月度总收入："
    ws_dash["H4"] = "=SUM(C4:C7)"
    ws_dash["G5"] = "月度总利润："
    ws_dash["H5"] = "=SUM(E4:E7)"
    ws_dash["G5"].font = Font(bold=True, color="FF0000")

    # --- 5. 图表分析 ---
    # 1. 利润构成柱状图
    bar = BarChart()
    bar.title = "各商品盈利能力对比"
    bar.y_axis.title = "金额 (元)"
    data = Reference(ws_dash, min_col=3, max_col=5, min_row=3, max_row=7)
    cats = Reference(ws_dash, min_col=1, min_row=4, max_row=7)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws_dash.add_chart(bar, "A10")

    # 2. 销售波动折线图 (直接从每日销量页取数据)
    line = LineChart()
    line.title = "每日销量波动趋势"
    line.y_axis.title = "销售件数"
    # 这里为了演示，直接引用 Sheet2 的前几行
    l_data = Reference(ws_qty, min_col=3, min_row=1, max_row=9)
    l_cats = Reference(ws_qty, min_col=1, min_row=2, max_row=9)
    line.add_data(l_data, titles_from_data=True)
    line.set_categories(l_cats)
    ws_dash.add_chart(line, "I10")

    # 保存
    wb.save(filename)
    print(f"深度分析报表 '{filename}' 已生成！")


generate_supermarket_pro("data/超市进销存深度分析表.xlsx")
