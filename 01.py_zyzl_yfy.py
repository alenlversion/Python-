import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
import os

# ================= 🛠️ 配置区域 =================

# 1. 文件路径
file_his = 'data/中医治疗(1).xlsx'
file_xt = 'data/附院中医非药物疗法字典(1).xlsx'

# 2. Sheet 名称
sheet_his = '住院统计'
sheet_xt = 'Sheet1'

# 3. 关键列名映射 (用于比对的核心列)
# 历史表列名
col_his_name = '项目名称'
# 字典表列名
col_xt_name = 'FareItemName'

# 4. 输出配置
output_file = 'data/双向差异对比报告_全量住院版.xlsx'

# =============================================================

print(f"🚀 开始执行双向差异比对...")

try:
    # 1. 读取数据
    print("⏳ 读取文件中...")
    df_his_raw = pd.read_excel(file_his, sheet_name=sheet_his)
    df_xt_raw = pd.read_excel(file_xt, sheet_name=sheet_xt)

    # 2. 数据标准化 (重命名列以便合并)
    # 我们只关心：项目代码(如果有), 项目名称, 来源

    # --- 处理历史表 ---
    # 假设历史表也有代码列，如果没有，用项目名称填充或留空，这里尝试自动查找包含 'Code' 或 '代码' 的列
    col_his_code = None
    for c in df_his_raw.columns:
        if 'Code' in c or '代码' in c:
            col_his_code = c
            break

    df_his = pd.DataFrame()
    df_his['FareItemName'] = df_his_raw[col_his_name].astype(str).str.strip()
    df_his['FareItemCode'] = df_his_raw[col_his_code].astype(str).str.strip() if col_his_code else ''
    df_his['Source'] = '历史表'
    df_his['OriginalRow'] = df_his_raw.index + 2  # 记录原行号方便追溯

    # --- 处理字典表 ---
    col_xt_code = None
    for c in df_xt_raw.columns:
        if 'Code' in c or '代码' in c:
            col_xt_code = c
            break

    df_xt = pd.DataFrame()
    df_xt['FareItemName'] = df_xt_raw[col_xt_name].astype(str).str.strip()
    df_xt['FareItemCode'] = df_xt_raw[col_xt_code].astype(str).str.strip() if col_xt_code else ''
    df_xt['Source'] = '字典表'
    df_xt['OriginalRow'] = df_xt_raw.index + 2

    # 3. 清洗空值
    df_his = df_his[df_his['FareItemName'] != 'nan']
    df_his = df_his[df_his['FareItemName'] != '']

    df_xt = df_xt[df_xt['FareItemName'] != 'nan']
    df_xt = df_xt[df_xt['FareItemName'] != '']

    print(f"   - 历史表有效项目数：{len(df_his)}")
    print(f"   - 字典表有效项目数：{len(df_xt)}")

    # 4. 计算差异集合
    set_his = set(df_his['FareItemName'].unique())
    set_xt = set(df_xt['FareItemName'].unique())

    only_in_his = set_his - set_xt  # 仅历史表有
    only_in_xt = set_xt - set_his  # 仅字典表有
    common = set_his & set_xt  # 两者都有

    print(f"\n📊 差异分析:")
    print(f"   - 仅存在于历史表：{len(only_in_his)}")
    print(f"   - 仅存在于字典表：{len(only_in_xt)}")
    print(f"   - 两者共有：{len(common)}")

    # 5. 构建结果 DataFrame
    result_data = []

    # A. 添加仅存在于历史表的数据
    for name in only_in_his:
        # 取第一条匹配记录作为代表，或者你可以选择合并所有代码
        row = df_his[df_his['FareItemName'] == name].iloc[0]
        result_data.append({
            'FareItemCode': row['FareItemCode'],
            'FareItemName': name,
            'Status': '❌ 仅存在于历史表 (字典表缺失)',
            'ColorType': 'orange'  # 自定义标记用于后续上色
        })

    # B. 添加仅存在于字典表的数据
    for name in only_in_xt:
        row = df_xt[df_xt['FareItemName'] == name].iloc[0]
        result_data.append({
            'FareItemCode': row['FareItemCode'],
            'FareItemName': name,
            'Status': '❌ 仅存在于字典表 (历史表缺失)',
            'ColorType': 'red'
        })

    # C. (可选) 添加共有数据，如果不需要可以注释掉这块
    # 为了报告清晰，通常差异报告只展示差异。如果你需要全量，取消下面注释
    """
    for name in common:
        row = df_xt[df_xt['FareItemName'] == name].iloc[0]
        result_data.append({
            'FareItemCode': row['FareItemCode'],
            'FareItemName': name,
            'Status': '✅ 匹配成功',
            'ColorType': 'green'
        })
    """

    df_result = pd.DataFrame(result_data)

    if df_result.empty:
        print("\n🎉 恭喜！两个表的项目名称完全一致（基于当前过滤条件）。")
        # 创建一个空表提示
        df_result = pd.DataFrame(
            [{'FareItemCode': '-', 'FareItemName': '无差异', 'Status': '完全一致', 'ColorType': 'green'}])

    # 调整列顺序：No_FareItem (序号), FareItemCode, FareItemName, 比对结果说明
    # 注意：Pandas 导出时会自动生成索引，我们可以重置索引作为 No_FareItem
    df_result.reset_index(drop=True, inplace=True)
    df_result.insert(0, 'No_FareItem', df_result.index + 1)

    # 重命名 Status 列为中文
    df_result.rename(columns={'Status': '比对结果说明'}, inplace=True)

    # 最终列顺序
    final_columns = ['No_FareItem', 'FareItemCode', 'FareItemName', '比对结果说明']
    df_final = df_result[final_columns]

    # 6. 写入 Excel 并应用颜色
    print(f"\n⏳ 正在生成报告并标色...")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name='差异详情', index=False)

        workbook = writer.book
        worksheet = writer.sheets['差异详情']

        # 定义样式
        fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        font_red = Font(color='9C0006', bold=True)

        fill_orange = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        font_orange = Font(color='9C6500', bold=True)

        fill_green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        font_green = Font(color='006100')

        # 表头加粗
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9D9D9', fill_type='solid')

        # 遍历行应用颜色
        # 数据从第2行开始 (index 0 -> row 2)
        for idx, row in df_result.iterrows():
            r_num = idx + 2
            color_type = row['ColorType']

            # 获取整行单元格 (从第1列到第4列)
            for col_idx in range(1, 5):
                cell = worksheet.cell(row=r_num, column=col_idx)

                if color_type == 'red':
                    cell.fill = fill_red
                    cell.font = font_red
                elif color_type == 'orange':
                    cell.fill = fill_orange
                    cell.font = font_orange
                elif color_type == 'green':
                    cell.fill = fill_green
                    cell.font = font_green

            # 自动调整行高
            worksheet.row_dimensions[r_num].height = 15

        # 调整列宽
        worksheet.column_dimensions['A'].width = 5  # 序号
        worksheet.column_dimensions['B'].width = 15  # 代码
        worksheet.column_dimensions['C'].width = 30  # 名称
        worksheet.column_dimensions['D'].width = 25  # 说明

    print(f"✅ 完成！")
    print(f"💾 文件已保存：{output_file}")
    print(f"💡 提示：")
    print(f"   - 🔴 红色行：字典表有，但历史表没有 (需确认是否漏录)")
    print(f"   - 🟠 橙色行：历史表有，但字典表没有 (需确认是否 obsolete 或新项)")
    # if green included: print(f"   - 🟢 绿色行：两边都有，名称一致")

except Exception as e:
    print(f"\n❌ 发生错误：{e}")
    import traceback

    traceback.print_exc()